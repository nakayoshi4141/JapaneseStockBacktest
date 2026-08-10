"""Japanese Stock Backtest - Excel output."""
from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import Config


class ExcelWriter:
    def __init__(self, output_file: Path | None = None) -> None:
        Config.create_output_directory()
        self.output_file = output_file or (Config.OUTPUT_DIR / Config.RESULT_FILE)

    def save(self, trade_history: pd.DataFrame, asset_history: pd.DataFrame, statistics: pd.DataFrame, yearly_result: pd.DataFrame) -> Path:
        with pd.ExcelWriter(self.output_file, engine="openpyxl") as writer:
            trade_history.to_excel(writer, sheet_name="TradeHistory", index=False)
            asset_history.to_excel(writer, sheet_name="AssetHistory", index=False)
            statistics.to_excel(writer, sheet_name="Statistics", index=False)
            yearly_result.to_excel(writer, sheet_name="YearlyResult", index=False)
        self.format_excel()
        return self.output_file

    def format_excel(self) -> None:
        wb = load_workbook(self.output_file)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for column_cells in ws.columns:
                max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
                letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[letter].width = min(max(max_length + 3, 10), 40)
        wb.save(self.output_file)

    def export(self, trade_history: pd.DataFrame, asset_history: pd.DataFrame, statistics: pd.DataFrame, yearly_result: pd.DataFrame) -> Path:
        return self.save(trade_history, asset_history, statistics, yearly_result)

    def export_trade_history_csv(self, trade_history: pd.DataFrame) -> Path:
        path = Config.OUTPUT_DIR / Config.TRADE_HISTORY_FILE
        trade_history.to_csv(path, index=False, encoding="utf-8-sig")
        return path
